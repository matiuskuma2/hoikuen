"""
児童利用予定表パーサー — v5.0 Phase B-4
1ファイル=複数園児対応（全シート読み込み）

実ファイル構造（2026年版）:
  D6  = 園児氏名
  F1  = 年 (col 6), J1 = 月 (col 10)
  左半分(日1-15):
    B12:B26 = 日付(datetime型), C=曜日(serial)
    D = 登所(time型), E = 降所(time型)
    F = 昼食, G = 朝おやつ, H = 午後おやつ, I = 夕食
  右半分(日16-31):
    J12:J27 = 日付(datetime型), K=曜日(serial)
    L = 登所(time型), M = 降所(time型)
    N = 昼食, O = 朝おやつ, P = 午後おやつ, Q = 夕食

v5.0 変更:
  - 実ファイル構造に完全対応（列マッピング修正）
  - 年=F1, 月=J1 に修正（旧: 年=J1, 月=M1）
  - 日付列がdatetime型の場合にday抽出対応
  - 食事4列（昼食・朝おやつ・午後おやつ・夕食）を個別読み取り
  - おやつ列の推定ロジック不要化（実ファイルは朝/午後おやつが別列）
  - 旧フォーマット（B6=名前, J1=年, M1=月）も自動検出しフォールバック
"""

import os
import re
from datetime import datetime, time as dt_time
from openpyxl import load_workbook
from engine.name_matcher import normalize_name


def parse_schedule_plans(file_path: str, target_year: int, target_month: int):
    """
    Parse schedule plan Excel file — supports MULTI-SHEET (1 file = multiple children).
    Each sheet is treated as one child's schedule.

    Returns:
        (results, warnings)

        results: list of (plans_dict, child_name) tuples
        warnings: list[dict]
    """
    warnings = []
    results = []  # list of (plans, child_name)
    filename = os.path.basename(file_path)

    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        warnings.append({
            "level": "error",
            "child_name": None,
            "message": f"予定表を開けません: {filename} ({e})",
            "suggestion": "ファイル形式が正しいか確認してください",
            "file": filename,
        })
        return results, warnings

    # Determine which sheets to process
    sheets_to_process = []
    if "原本" in wb.sheetnames:
        # If "原本" exists, process only that sheet (legacy 1-sheet mode)
        sheets_to_process = [wb["原本"]]
    else:
        # Process ALL sheets — each sheet = 1 child
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Skip obviously empty or meta sheets
            if ws.max_row is not None and ws.max_row < 5:
                continue
            sheets_to_process.append(ws)

    if len(sheets_to_process) == 0:
        warnings.append({
            "level": "error",
            "child_name": None,
            "message": f"「{filename}」: 処理可能なシートがありません",
            "suggestion": "予定表のシート構成を確認してください",
            "file": filename,
        })
        wb.close()
        return results, warnings

    if len(sheets_to_process) > 1:
        warnings.append({
            "level": "info",
            "child_name": None,
            "message": f"「{filename}」: {len(sheets_to_process)}シート検出 → 全シートを読み込みます",
            "suggestion": None,
            "file": filename,
        })

    for ws in sheets_to_process:
        sheet_label = f"{filename}[{ws.title}]"
        plans, child_name, sheet_warnings = _parse_single_sheet(
            ws, filename, sheet_label, target_year, target_month
        )
        warnings.extend(sheet_warnings)
        if child_name:
            results.append((plans, child_name))

    wb.close()
    return results, warnings


def _detect_layout(cells):
    """
    Detect the sheet layout format.
    Returns a layout dict with column mappings.

    新フォーマット (2026年版):
      年: F1(col6), 月: J1(col10)
      名前: D6(col4)
      左半分: B=日付, D=登所, E=降所, F=昼食, G=朝おやつ, H=午後おやつ, I=夕食
      右半分: J=日付, L=登所, M=降所, N=昼食, O=朝おやつ, P=午後おやつ, Q=夕食

    旧フォーマット:
      年: J1(col10), 月: M1(col13)
      名前: B6(col2)
      左半分: B=日付, D=登所, G=降所, J=昼食, K=おやつ, L=夕食
      右半分: M=日付, O=登所, R=降所, U=昼食, V=おやつ, W=夕食
    """
    # Check header row 11 for meal column layout
    # 新フォーマット: F11=昼食, G11=朝おやつ, H11=午後おやつ, I11=夕食
    # 旧フォーマット: J11=昼食, K11=おやつ, L11=夕食
    f11 = str(cells.get((11, 6), "")).strip()  # F11
    g11 = str(cells.get((11, 7), "")).strip()  # G11

    # Check if header text contains 昼食 at col F (new format) or at col J (old format)
    has_new_header = "昼食" in f11 or "おやつ" in g11

    # Also check: in the new format, year is at F1 (col 6) and a valid year
    f1_val = cells.get((1, 6))
    j1_val = cells.get((1, 10))

    # New format: F1 has year (2020-2030), J1 has month (1-12)
    f1_is_year = False
    j1_is_month = False
    if f1_val is not None:
        try:
            f1_int = int(f1_val)
            if 2020 <= f1_int <= 2030:
                f1_is_year = True
        except (ValueError, TypeError):
            pass
    if j1_val is not None:
        try:
            j1_int = int(j1_val)
            if 1 <= j1_int <= 12:
                j1_is_month = True
        except (ValueError, TypeError):
            pass

    is_new_format = has_new_header or (f1_is_year and j1_is_month)

    if is_new_format:
        return {
            "format": "new",
            "year_pos": (1, 6),       # F1
            "month_pos": (1, 10),     # J1
            "name_pos": (6, 4),       # D6
            # Left half (days 1-15) - data starts at row 12
            "left_date_col": 2,       # B
            "left_start_col": 4,      # D
            "left_end_col": 5,        # E
            "left_lunch_col": 6,      # F
            "left_am_snack_col": 7,   # G
            "left_pm_snack_col": 8,   # H
            "left_dinner_col": 9,     # I
            # Right half (days 16-31) - data starts at row 12
            "right_date_col": 10,     # J
            "right_start_col": 12,    # L
            "right_end_col": 13,      # M
            "right_lunch_col": 14,    # N
            "right_am_snack_col": 15, # O
            "right_pm_snack_col": 16, # P
            "right_dinner_col": 17,   # Q
            "has_separate_snacks": True,
        }
    else:
        return {
            "format": "legacy",
            "year_pos": (1, 10),      # J1
            "month_pos": (1, 13),     # M1
            "name_pos": (6, 2),       # B6
            # Left half
            "left_date_col": 2,       # B
            "left_start_col": 4,      # D
            "left_end_col": 7,        # G
            "left_lunch_col": 10,     # J
            "left_snack_col": 11,     # K
            "left_dinner_col": 12,    # L
            # Right half
            "right_date_col": 13,     # M
            "right_start_col": 15,    # O
            "right_end_col": 18,      # R
            "right_lunch_col": 21,    # U
            "right_snack_col": 22,    # V
            "right_dinner_col": 23,   # W
            "has_separate_snacks": False,
        }


def _parse_single_sheet(ws, filename: str, sheet_label: str,
                        target_year: int, target_month: int):
    """Parse a single worksheet as one child's schedule."""
    warnings = []
    plans = {}
    child_name = None

    # Read all cells into a dict for random access
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cells[(cell.row, cell.column)] = cell.value

    # Skip sheets that look empty (fewer than 3 non-empty cells)
    if len(cells) < 3:
        return plans, None, []

    # ── Detect layout format ──
    layout = _detect_layout(cells)

    # ── Extract child name ──
    name_pos = layout["name_pos"]
    child_name_raw = cells.get(name_pos)
    if child_name_raw:
        child_name = normalize_name(str(child_name_raw))

    if not child_name:
        # Fallback: try multiple positions
        fallback_positions = [(6, 2), (6, 3), (6, 4), (5, 2), (5, 4)]
        for fallback_pos in fallback_positions:
            if fallback_pos == name_pos:
                continue  # already tried
            fallback_val = cells.get(fallback_pos)
            if fallback_val and str(fallback_val).strip():
                # Skip if it looks like a label (e.g., "お子様名（")
                val_str = str(fallback_val).strip()
                if "お子様" in val_str or "（" in val_str or "申込" in val_str:
                    continue
                child_name = normalize_name(val_str)
                warnings.append({
                    "level": "info",
                    "child_name": child_name,
                    "message": f"「{sheet_label}」: {_cell_ref(name_pos)}が空のため{_cell_ref(fallback_pos)}から園児名を検出",
                    "suggestion": None,
                    "file": filename,
                })
                break

    if not child_name:
        # Try U10 (col 21, row 10) — some sheets have the name there
        u10_val = cells.get((10, 21))
        if u10_val and str(u10_val).strip():
            child_name = normalize_name(str(u10_val))
            warnings.append({
                "level": "info",
                "child_name": child_name,
                "message": f"「{sheet_label}」: U10から園児名を検出",
                "suggestion": None,
                "file": filename,
            })

    if not child_name:
        # Try sheet name as child name (common pattern: sheet name = child name)
        sheet_name_candidate = ws.title.strip()
        if sheet_name_candidate and len(sheet_name_candidate) >= 2 and not sheet_name_candidate.startswith("Sheet"):
            # Exclude common non-name sheet titles and pure numbers
            skip_titles = {"原本", "設定", "マスタ", "一覧", "集計", "sheet1", "sheet2", "sheet3"}
            if (sheet_name_candidate.lower() not in skip_titles
                    and not sheet_name_candidate.isdigit()):
                child_name = normalize_name(sheet_name_candidate)
                warnings.append({
                    "level": "warn",
                    "child_name": child_name,
                    "message": f"「{sheet_label}」: セルから園児名を検出できず、シート名「{ws.title}」から推測",
                    "suggestion": "予定表のD6セルに園児名を入力してください",
                    "file": filename,
                })

    if not child_name:
        # Last resort: extract from filename (only for single-sheet files)
        name_from_file = _extract_name_from_filename(filename)
        if name_from_file:
            child_name = name_from_file
            warnings.append({
                "level": "warn",
                "child_name": child_name,
                "message": f"「{sheet_label}」: セルから園児名を検出できず、ファイル名から推測",
                "suggestion": "予定表のD6セルに園児名を入力してください",
                "file": filename,
            })
        else:
            warnings.append({
                "level": "error",
                "child_name": None,
                "message": f"「{sheet_label}」: 園児名を検出できません",
                "suggestion": "D6セルに園児名が入力されているか確認してください",
                "file": filename,
            })
            return plans, child_name, warnings

    # ── Check year/month ──
    year_pos = layout["year_pos"]
    month_pos = layout["month_pos"]
    file_year = _safe_int(cells.get(year_pos))
    file_month = _safe_int(cells.get(month_pos))

    # Try alternate locations for year/month if primary fails
    if not file_year or not (2020 <= file_year <= 2030):
        for pos in [(1, 6), (1, 10), (1, 11), (8, 8), (8, 11)]:
            v = _safe_int(cells.get(pos))
            if v and 2020 <= v <= 2030:
                file_year = v
                break
    if not file_month or not (1 <= file_month <= 12):
        for pos in [(1, 10), (1, 13), (1, 14), (8, 11), (8, 14)]:
            v = _safe_int(cells.get(pos))
            if v and 1 <= v <= 12:
                # Avoid confusing year and month
                if v != file_year:
                    file_month = v
                    break

    month_mismatch = False
    if file_year and file_month:
        if file_year != target_year or file_month != target_month:
            month_mismatch = True
            warnings.append({
                "level": "warn",
                "child_name": child_name,
                "message": f"「{sheet_label}」の年月({file_year}年{file_month}月)が対象月({target_year}年{target_month}月)と不一致",
                "suggestion": "正しいファイルか確認してください。このファイルのデータは使用されますが、日付のずれにご注意ください",
                "file": filename,
            })
    elif not file_year or not file_month:
        warnings.append({
            "level": "info",
            "child_name": child_name,
            "message": f"「{sheet_label}」: 年月セルが空または読み取れません (year={file_year}, month={file_month})",
            "suggestion": None,
            "file": filename,
        })

    # ── Parse left half: days 1-15 ──
    for i in range(15):
        row = 12 + i
        day = _extract_day(cells.get((row, layout["left_date_col"])))
        if day is None or day < 1 or day > 31:
            continue

        if layout["has_separate_snacks"]:
            plan = _extract_plan_v2(cells, row,
                                    start_col=layout["left_start_col"],
                                    end_col=layout["left_end_col"],
                                    lunch_col=layout["left_lunch_col"],
                                    am_snack_col=layout["left_am_snack_col"],
                                    pm_snack_col=layout["left_pm_snack_col"],
                                    dinner_col=layout["left_dinner_col"])
        else:
            plan = _extract_plan_legacy(cells, row,
                                        start_col=layout["left_start_col"],
                                        end_col=layout["left_end_col"],
                                        lunch_col=layout["left_lunch_col"],
                                        snack_col=layout["left_snack_col"],
                                        dinner_col=layout["left_dinner_col"])
        if plan:
            plan["day"] = day
            plan["child_name"] = child_name
            plan["source_file"] = filename
            plan["month_mismatch"] = month_mismatch
            plans[day] = plan

    # ── Parse right half: days 16-31 ──
    for i in range(16):
        row = 12 + i
        day = _extract_day(cells.get((row, layout["right_date_col"])))
        if day is None or day < 1 or day > 31:
            continue

        if layout["has_separate_snacks"]:
            plan = _extract_plan_v2(cells, row,
                                    start_col=layout["right_start_col"],
                                    end_col=layout["right_end_col"],
                                    lunch_col=layout["right_lunch_col"],
                                    am_snack_col=layout["right_am_snack_col"],
                                    pm_snack_col=layout["right_pm_snack_col"],
                                    dinner_col=layout["right_dinner_col"])
        else:
            plan = _extract_plan_legacy(cells, row,
                                        start_col=layout["right_start_col"],
                                        end_col=layout["right_end_col"],
                                        lunch_col=layout["right_lunch_col"],
                                        snack_col=layout["right_snack_col"],
                                        dinner_col=layout["right_dinner_col"])
        if plan:
            plan["day"] = day
            plan["child_name"] = child_name
            plan["source_file"] = filename
            plan["month_mismatch"] = month_mismatch
            plans[day] = plan

    # ── Summary ──
    if len(plans) == 0:
        warnings.append({
            "level": "warn",
            "child_name": child_name,
            "message": f"「{sheet_label}」: 有効な利用予定が0件です",
            "suggestion": "予定表の内容が正しいか確認してください",
            "file": filename,
        })

    return plans, child_name, warnings


def parse_multiple_schedules(
    file_paths: list[tuple[str, str]],  # [(file_path, original_filename), ...]
    target_year: int,
    target_month: int,
) -> tuple[dict[str, dict], list[str], list[dict]]:
    """
    Parse multiple schedule files and aggregate results.
    Each file may contain multiple sheets (one child per sheet).

    Returns:
        (all_plans, child_names, all_warnings)

        all_plans: dict[child_name → {day: plan}]
        child_names: list of detected child names
        all_warnings: aggregated warnings
    """
    all_plans = {}
    child_names = []
    all_warnings = []

    for file_path, orig_name in file_paths:
        results, file_warnings = parse_schedule_plans(file_path, target_year, target_month)
        all_warnings.extend(file_warnings)

        for plans, child_name in results:
            if child_name:
                child_names.append(child_name)
                if child_name in all_plans:
                    # Duplicate schedule for same child
                    all_warnings.append({
                        "level": "warn",
                        "child_name": child_name,
                        "message": f"園児「{child_name}」の予定表が複数アップロードされています。後のデータで上書きします",
                        "suggestion": None,
                        "file": orig_name,
                    })
                all_plans[child_name] = plans

    return all_plans, child_names, all_warnings


def _extract_day(val) -> int | None:
    """Extract day number from a cell value.
    Handles: int, datetime, float, string."""
    if val is None:
        return None

    # datetime型: 2026-01-05 → day=5
    if isinstance(val, datetime):
        return val.day

    # time型はスキップ（日付ではない）
    if isinstance(val, dt_time):
        return None

    # int型: そのまま
    if isinstance(val, int):
        if 1 <= val <= 31:
            return val
        return None

    # float型: 小数点以下なしなら整数として扱う
    if isinstance(val, float):
        if val == int(val) and 1 <= int(val) <= 31:
            return int(val)
        return None

    # 文字列
    s = str(val).strip()
    try:
        d = int(s)
        if 1 <= d <= 31:
            return d
    except ValueError:
        pass

    return None


def _extract_plan_v2(cells, row, start_col, end_col,
                     lunch_col, am_snack_col, pm_snack_col, dinner_col):
    """Extract a single day's plan from cell values (new format with separate snack columns)."""
    start = _parse_time_cell(cells.get((row, start_col)))
    end = _parse_time_cell(cells.get((row, end_col)))

    # If no times at all, skip this day
    if start is None and end is None:
        return None

    lunch = _is_flag(cells.get((row, lunch_col)))
    am_snack = _is_flag(cells.get((row, am_snack_col)))
    pm_snack = _is_flag(cells.get((row, pm_snack_col)))
    dinner = _is_flag(cells.get((row, dinner_col)))

    return {
        "planned_start": start,
        "planned_end": end,
        "lunch_flag": 1 if lunch else 0,
        "am_snack_flag": 1 if am_snack else 0,
        "pm_snack_flag": 1 if pm_snack else 0,
        "dinner_flag": 1 if dinner else 0,
    }


def _extract_plan_legacy(cells, row, start_col, end_col,
                         lunch_col, snack_col, dinner_col):
    """Extract a single day's plan from cell values (legacy format with single snack column).
    Uses time-based inference to split snack into morning/afternoon."""
    start = _parse_time_cell(cells.get((row, start_col)))
    end = _parse_time_cell(cells.get((row, end_col)))

    # If no times at all, skip this day
    if start is None and end is None:
        return None

    lunch = _is_flag(cells.get((row, lunch_col)))
    snack = _is_flag(cells.get((row, snack_col)))
    dinner = _is_flag(cells.get((row, dinner_col)))

    # ★ 旧フォーマットの「おやつ」列(K/V)は1列のみ。
    # 朝おやつ・午後おやつは予定時間帯から推定:
    am_snack = False
    pm_snack = False
    if snack:
        start_min = _time_to_minutes(start) if start else None
        end_min = _time_to_minutes(end) if end else None
        if start_min is not None and start_min <= 600:
            am_snack = True
        if end_min is not None and end_min >= 900:
            pm_snack = True
        if not am_snack and not pm_snack:
            pm_snack = True

    return {
        "planned_start": start,
        "planned_end": end,
        "lunch_flag": 1 if lunch else 0,
        "am_snack_flag": 1 if am_snack else 0,
        "pm_snack_flag": 1 if pm_snack else 0,
        "dinner_flag": 1 if dinner else 0,
    }


def _parse_time_cell(val) -> str | None:
    """Parse a time cell value to HH:MM format"""
    if val is None:
        return None

    if isinstance(val, dt_time):
        return val.strftime("%H:%M")

    if isinstance(val, datetime):
        # datetime with only time part (e.g., 1900-01-01 09:00:00 from Excel)
        return val.strftime("%H:%M")

    if isinstance(val, (int, float)):
        if 0 <= val < 1:
            total_min = round(val * 24 * 60)
            h = total_min // 60
            m = total_min % 60
            return f"{h:02d}:{m:02d}"
        # Might be hours (e.g., 9 → 09:00)
        if 0 < val < 24 and val == int(val):
            return f"{int(val):02d}:00"
        return None

    s = str(val).strip()
    if not s:
        return None

    match = re.match(r'(\d{1,2}):(\d{2})', s)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"

    # Try just a number like "900" → "9:00"
    match = re.match(r'^(\d{1,2})(\d{2})$', s)
    if match:
        h = int(match.group(1))
        m = int(match.group(2))
        if 0 <= h <= 23 and 0 <= m <= 59:
            return f"{h:02d}:{m:02d}"

    return None


def _time_to_minutes(time_str: str | None) -> int | None:
    """HH:MM → total minutes (for snack time inference)"""
    if not time_str:
        return None
    parts = time_str.split(":")
    if len(parts) >= 2:
        try:
            return int(parts[0]) * 60 + int(parts[1])
        except (ValueError, TypeError):
            return None
    return None


def _is_flag(val) -> bool:
    """Check if a cell indicates a meal flag (〇, ○, 1, TRUE, etc)"""
    if val is None:
        return False
    s = str(val).strip()
    return s in ('〇', '○', 'O', 'o', '1', 'TRUE', 'true', '◯', '●', '✓', '✔')


def _safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _cell_ref(pos: tuple) -> str:
    """Convert (row, col) to Excel cell reference like 'B6'"""
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(pos[1])}{pos[0]}"


def _extract_name_from_filename(filename: str) -> str | None:
    """Try to extract child name from filename like '田中太郎_予定表.xlsx'"""
    # Remove extension
    name = os.path.splitext(filename)[0]

    # Remove common suffixes
    for suffix in ["_予定表", "_予定", "_利用予定", "予定表", "利用予定表", "_schedule",
                   "児童利用予定表", "児童利用"]:
        name = name.replace(suffix, "")

    # Remove year/month patterns
    name = re.sub(r'\d{4}[年/\-]?\d{1,2}[月]?', '', name)
    name = re.sub(r'\d{1,2}月', '', name)

    # Remove leading/trailing numbers and spaces
    name = re.sub(r'^\d+', '', name)
    name = re.sub(r'\(\d+\)$', '', name)  # Remove (1), (2) etc.

    # Clean up
    name = name.strip("_- 　()")

    if len(name) >= 2:
        return normalize_name(name)
    return None
