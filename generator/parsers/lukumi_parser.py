"""
ルクミー登降園データパーサー — v3.1 Phase B-1
Excel(.xlsx) or CSV から attendance_records + children基本情報を抽出

ルクミーの列構造 (Excel/CSV共通):
  A=クラス名, B=園児姓, C=園児名, D=日付,
  E=登園日時, F=降園日時, G=メモ, H=園児ID,
  I=姓よみ, J=名よみ, K=生年月日, L=クラス年齢

ヘッダー列名（自動検出用）:
  "クラス名", "園児姓"/"姓", "園児名"/"名",
  "日付"/"登降園日", "登園日時"/"登園", "降園日時"/"降園",
  "メモ", "園児ID"/"子どもID", "姓よみ", "名よみ",
  "生年月日", "クラス年齢"/"年齢"

ルクミーID = SSOT (最も安定な突合キー)
"""

import os
import re
import csv
from datetime import datetime, date, time as dt_time
from openpyxl import load_workbook
from engine.name_matcher import normalize_name


# ── Column name patterns for auto-detection ──
_COL_PATTERNS = {
    "class_name":  ["クラス名", "クラス", "class"],
    "surname":     ["園児姓", "姓", "苗字"],
    "firstname":   ["園児名", "名前"],
    "date":        ["日付", "登降園日", "date"],
    "checkin":     ["登園日時", "登園", "checkin", "check-in"],
    "checkout":    ["降園日時", "降園", "checkout", "check-out"],
    "memo":        ["メモ", "備考", "memo"],
    "lukumi_id":   ["園児ID", "子どもID", "児童ID", "id"],
    "kana_sei":    ["姓よみ", "姓読み", "セイ"],
    "kana_mei":    ["名よみ", "名読み", "メイ"],
    "birth_date":  ["生年月日", "誕生日", "birthday"],
    "age_class":   ["クラス年齢", "年齢", "歳児", "age"],
}

# ★ "名" 単独は "クラス名" や "姓名" にも含まれるため、
#   部分一致 (pattern in cell_str) では誤検出が起きる。
#   _detect_columns で「完全一致を優先」するロジックで対応。


def parse_lukumi(file_path: str, target_year: int, target_month: int) -> tuple[list[dict], list[dict], list[dict]]:
    """
    Parse Lukumi attendance file (Excel or CSV).

    Returns:
        (attendance_records, children_info, warnings)

        attendance_records: list of {
            lukumi_id, name, year, month, day,
            actual_checkin, actual_checkout, memo, class_name
        }

        children_info: list of {
            lukumi_id, name, name_kana, birth_date,
            age_class, class_name, enrollment_type
        }

        warnings: list of {level, child_name, message, suggestion}
    """
    ext = os.path.splitext(file_path)[1].lower()
    warnings = []

    if ext == '.csv':
        attendance, children, w = _parse_csv(file_path, target_year, target_month)
    else:
        attendance, children, w = _parse_xlsx(file_path, target_year, target_month)

    warnings.extend(w)

    # ── Validation ──
    if len(attendance) == 0:
        warnings.append({
            "level": "error",
            "child_name": None,
            "message": f"ルクミーデータから{target_year}年{target_month}月の出席レコードが0件です",
            "suggestion": "ファイルの期間を確認してください",
        })

    if len(children) == 0:
        warnings.append({
            "level": "error",
            "child_name": None,
            "message": "ルクミーデータから園児情報を抽出できませんでした",
            "suggestion": "ファイルのフォーマットを確認してください",
        })

    # Validate: check for children with checkin but no checkout
    checkin_only = {}
    for a in attendance:
        key = (a["lukumi_id"], a["day"])
        if a["actual_checkin"] and not a["actual_checkout"]:
            checkin_only[key] = a["name"]
    if checkin_only:
        for (lid, day), name in list(checkin_only.items())[:5]:  # limit warnings
            warnings.append({
                "level": "warn",
                "child_name": name,
                "message": f"{target_month}月{day}日: 登園記録あり・降園記録なし",
                "suggestion": "ルクミーの記録漏れの可能性があります",
            })
        if len(checkin_only) > 5:
            warnings.append({
                "level": "warn",
                "child_name": None,
                "message": f"他に{len(checkin_only) - 5}件の降園未記録があります",
                "suggestion": None,
            })

    return attendance, children, warnings


def _parse_xlsx(file_path: str, target_year: int, target_month: int):
    warnings = []
    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        return [], [], [{
            "level": "error", "child_name": None,
            "message": f"ルクミーExcelを開けません: {e}",
            "suggestion": "ファイルが破損していないか確認してください",
        }]

    ws = wb.active

    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], [], [{
            "level": "error", "child_name": None,
            "message": "ルクミーファイルにデータ行がありません",
            "suggestion": "正しいファイルか確認してください",
        }]

    # Auto-detect header row and column mapping
    col_map, header_row_idx, hw = _detect_columns(rows)
    warnings.extend(hw)

    if col_map is None:
        return [], [], warnings

    # Parse data rows
    attendance = []
    children_map = {}

    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]
        rec, child_info = _parse_row(row, col_map, target_year, target_month)
        if rec:
            attendance.append(rec)
        if child_info and child_info["lukumi_id"] not in children_map:
            children_map[child_info["lukumi_id"]] = child_info

    return attendance, list(children_map.values()), warnings


def _parse_csv(file_path: str, target_year: int, target_month: int):
    warnings = []

    # Try different encodings
    rows = None
    for encoding in ['utf-8-sig', 'utf-8', 'shift_jis', 'cp932']:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                rows = [tuple(row) for row in reader]
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if not rows or len(rows) < 2:
        return [], [], [{
            "level": "error", "child_name": None,
            "message": "ルクミーCSVを読み込めません",
            "suggestion": "UTF-8またはShift_JIS形式であることを確認してください",
        }]

    # Auto-detect columns
    col_map, header_row_idx, hw = _detect_columns(rows)
    warnings.extend(hw)

    if col_map is None:
        return [], [], warnings

    attendance = []
    children_map = {}

    for row_idx in range(header_row_idx + 1, len(rows)):
        row = rows[row_idx]
        rec, child_info = _parse_row(row, col_map, target_year, target_month)
        if rec:
            attendance.append(rec)
        if child_info and child_info["lukumi_id"] not in children_map:
            children_map[child_info["lukumi_id"]] = child_info

    return attendance, list(children_map.values()), warnings


def _detect_columns(rows: list[tuple]) -> tuple[dict | None, int, list[dict]]:
    """
    Auto-detect column mapping from header row.
    
    ★ 検出ロジック:
      Pass 1: 完全一致 (cell_str == pattern) — 最も信頼性が高い
      Pass 2: 部分一致 (pattern in cell_str) — 曖昧だが fallback
      ※ 一度マッピングされた列は後のパスで上書きしない
      ※ "名" 単独パターンは "クラス名" にも含まれるため、
         firstname は "園児名", "名前" のみで検出。fallback で列 C (index 2) を仮定。
    
    Returns (col_map, header_row_index, warnings)
    """
    warnings = []

    # Check first 5 rows for header
    for header_idx in range(min(5, len(rows))):
        header = rows[header_idx]
        col_map = {}
        used_cols = set()  # Track which column indices are already assigned

        # ── Pass 1: 完全一致 ──
        for col_idx, cell_val in enumerate(header):
            if cell_val is None:
                continue
            cell_str = str(cell_val).strip()

            for field, patterns in _COL_PATTERNS.items():
                if field in col_map:
                    continue  # Already mapped
                for pattern in patterns:
                    if pattern.lower() == cell_str.lower():
                        col_map[field] = col_idx
                        used_cols.add(col_idx)
                        break

        # ── Pass 2: 部分一致 (未マッチのフィールドのみ) ──
        for col_idx, cell_val in enumerate(header):
            if cell_val is None or col_idx in used_cols:
                continue
            cell_str = str(cell_val).strip()

            for field, patterns in _COL_PATTERNS.items():
                if field in col_map:
                    continue
                for pattern in patterns:
                    if pattern in cell_str and len(pattern) >= 2:
                        # ★ 2文字以上のパターンのみ部分一致を許可
                        # "名" (1文字) は誤検出しやすいので除外
                        col_map[field] = col_idx
                        used_cols.add(col_idx)
                        break

        # ── Fallback: firstname が未検出なら surname の隣 (右1列) を仮定 ──
        if "firstname" not in col_map and "surname" in col_map:
            next_col = col_map["surname"] + 1
            if next_col < len(header) and next_col not in used_cols:
                col_map["firstname"] = next_col
                used_cols.add(next_col)
                warnings.append({
                    "level": "info", "child_name": None,
                    "message": f"園児名列を自動検出できず。姓列の右隣(列{next_col + 1})を名列と仮定します",
                    "suggestion": None,
                })

        # Check if we have minimum required columns
        required = {"surname", "date", "lukumi_id"}
        if required.issubset(col_map.keys()):
            if "checkin" not in col_map:
                warnings.append({
                    "level": "warn", "child_name": None,
                    "message": "登園日時列が自動検出できません。デフォルト位置(E列)を使用します",
                    "suggestion": None,
                })
                col_map["checkin"] = 4

            if "checkout" not in col_map:
                warnings.append({
                    "level": "warn", "child_name": None,
                    "message": "降園日時列が自動検出できません。デフォルト位置(F列)を使用します",
                    "suggestion": None,
                })
                col_map["checkout"] = 5

            return col_map, header_idx, warnings

    # Fallback: assume standard Lukumi column order
    warnings.append({
        "level": "warn", "child_name": None,
        "message": "ヘッダー自動検出できず。ルクミー標準列順（A=クラス名, B=姓, ...）を仮定します",
        "suggestion": None,
    })

    col_map = {
        "class_name": 0, "surname": 1, "firstname": 2, "date": 3,
        "checkin": 4, "checkout": 5, "memo": 6, "lukumi_id": 7,
        "kana_sei": 8, "kana_mei": 9, "birth_date": 10, "age_class": 11,
    }
    return col_map, 0, warnings


def _parse_row(row: tuple, col_map: dict, target_year: int, target_month: int):
    """Parse a single data row. Returns (attendance_record | None, child_info | None)"""
    if len(row) < max(col_map.values()) + 1:
        return None, None

    def get(field):
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    surname = str(get("surname") or "").strip()
    firstname = str(get("firstname") or "").strip() if "firstname" in col_map else ""
    lukumi_id = str(get("lukumi_id") or "").strip()

    full_name = f"{surname} {firstname}".strip() if firstname else surname
    if not full_name or not lukumi_id:
        return None, None

    # Parse date
    dt = _parse_date(get("date"))
    if dt is None:
        return None, None

    # Filter by target month
    if dt.year != target_year or dt.month != target_month:
        return None, None

    # Parse times
    checkin = _parse_time(get("checkin"))
    checkout = _parse_time(get("checkout"))

    rec = {
        "lukumi_id": lukumi_id,
        "name": full_name,
        "year": dt.year,
        "month": dt.month,
        "day": dt.day,
        "actual_checkin": checkin,
        "actual_checkout": checkout,
        "memo": str(get("memo") or "").strip() if get("memo") else None,
        "class_name": str(get("class_name") or "").strip(),
    }

    # Child info
    kana_sei = str(get("kana_sei") or "").strip() if get("kana_sei") else ""
    kana_mei = str(get("kana_mei") or "").strip() if get("kana_mei") else ""
    name_kana = f"{kana_sei} {kana_mei}".strip() or None

    birth_date = None
    bd = get("birth_date")
    if bd:
        bd_parsed = _parse_date(bd)
        if bd_parsed:
            birth_date = bd_parsed.strftime("%Y-%m-%d")

    age_class = None
    ac = get("age_class")
    if ac is not None:
        try:
            age_class = int(ac)
        except (ValueError, TypeError):
            m = re.match(r'(\d+)', str(ac))
            if m:
                age_class = int(m.group(1))

    child_info = {
        "lukumi_id": lukumi_id,
        "name": full_name,
        "name_kana": name_kana,
        "birth_date": birth_date,
        "age_class": age_class,
        "class_name": str(get("class_name") or "").strip(),
        "enrollment_type": "月極",  # Default; override from roster
    }

    return rec, child_info


# ── Date/Time Parsing Utilities ──

def _parse_date(val) -> date | None:
    """Parse various date formats"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    s = str(val).strip()
    if not s:
        return None

    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S",
                "%Y-%m-%dT%H:%M:%S", "%Y年%m月%d日", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_time(val) -> str | None:
    """Parse time value to HH:MM format"""
    if val is None:
        return None

    if isinstance(val, dt_time):
        return val.strftime("%H:%M")

    if isinstance(val, datetime):
        return val.strftime("%H:%M")

    if isinstance(val, (int, float)):
        # Excel time serial (0.0-1.0)
        if 0 <= val < 1:
            total_min = round(val * 24 * 60)
            h = total_min // 60
            m = total_min % 60
            return f"{h:02d}:{m:02d}"
        return None

    s = str(val).strip()
    if not s:
        return None

    # HH:MM:SS or HH:MM
    match = re.match(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', s)
    if match:
        h = int(match.group(1))
        m = int(match.group(2))
        return f"{h:02d}:{m:02d}"

    # Extract from datetime string "2026/01/05 08:30:00"
    match = re.search(r'(\d{2}):(\d{2})(?::(\d{2}))?', s)
    if match:
        h = int(match.group(1))
        m = int(match.group(2))
        return f"{h:02d}:{m:02d}"

    return None
