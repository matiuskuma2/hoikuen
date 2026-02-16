"""
園児名簿パーサー — v3.1 Phase B-2
日報テンプレート内の◆園児名簿シートから園児マスタを抽出

目的:
  - ルクミーの基本情報を名簿情報で補完（歳児クラス、利用区分、第何子、アレルギー等）
  - テンプレート内の園児行位置を特定（後続のwriter用）

シート探索順:
  1. "◆園児名簿" (完全一致)
  2. "園児名簿" (部分一致)
  3. シート名に "名簿" を含むもの

列構造（自動検出 + フォールバック）:
  B列=No, C列=クラス名, D列=氏名, E列=利用区分, F列=第何子,
  G列=生年月日, H列=歳児, I列=アレルギー, J列=徴収方法

設計方針:
  - ヘッダー行を自動検出（"No", "氏名" 等のキーワード）
  - 空行で打ち切り（名簿末尾検出）
  - 全フィールド optional（取れるものだけ取る）
"""

import re
from datetime import datetime, date
from openpyxl import load_workbook
from engine.name_matcher import normalize_name


# Column header patterns for auto-detection
_ROSTER_COL_PATTERNS = {
    "no":              ["No", "NO", "no", "番号", "#"],
    "class_name":      ["クラス", "クラス名", "組"],
    "name":            ["氏名", "園児名", "名前", "児童名"],
    "enrollment_type": ["利用区分", "利用種別", "区分", "種別"],
    "child_order":     ["第何子", "何子", "きょうだい", "兄弟"],
    "birth_date":      ["生年月日", "誕生日", "birthday"],
    "age_class":       ["歳児", "年齢", "クラス年齢"],
    "is_allergy":      ["アレルギー", "allergy"],
    "collection":      ["徴収方法", "徴収", "支払", "口座"],
}


def parse_roster(template_path: str) -> tuple[list[dict], list[dict]]:
    """
    Parse child roster from daily report template.

    Returns:
        (children, warnings)
        
        children: list of {
            roster_no, name, name_norm, class_name, enrollment_type,
            child_order, birth_date, age_class, is_allergy, collection_method,
            row_index (Excel row position in sheet)
        }
        
        warnings: list of {level, child_name, message, suggestion}
    """
    children = []
    warnings = []

    try:
        wb = load_workbook(template_path, data_only=True, read_only=False)
    except Exception as e:
        return children, [{
            "level": "error", "child_name": None,
            "message": f"テンプレートを開けません: {e}",
            "suggestion": "ファイルが破損していないか確認してください",
        }]

    # ── Find roster sheet ──
    roster_sheet = None
    for priority_name in ["◆園児名簿", "園児名簿"]:
        for sheet_name in wb.sheetnames:
            if priority_name in sheet_name:
                roster_sheet = wb[sheet_name]
                break
        if roster_sheet:
            break

    if not roster_sheet:
        for sheet_name in wb.sheetnames:
            if "名簿" in sheet_name:
                roster_sheet = wb[sheet_name]
                break

    if not roster_sheet:
        wb.close()
        return children, [{
            "level": "warn", "child_name": None,
            "message": "園児名簿シートが見つかりません",
            "suggestion": "テンプレートに「園児名簿」シートがあるか確認してください",
        }]

    # ── Read all rows into list ──
    all_rows = []
    for row in roster_sheet.iter_rows(values_only=False):
        row_data = []
        for cell in row:
            row_data.append({
                "value": cell.value,
                "row": cell.row,
                "col": cell.column,
            })
        all_rows.append(row_data)

    wb.close()

    if len(all_rows) < 2:
        return children, [{
            "level": "warn", "child_name": None,
            "message": "園児名簿シートにデータがありません",
            "suggestion": None,
        }]

    # ── Auto-detect header row and columns ──
    col_map = None
    header_row_idx = None

    for row_idx in range(min(10, len(all_rows))):
        row = all_rows[row_idx]
        detected = {}

        for cell_info in row:
            if cell_info["value"] is None:
                continue
            cell_str = str(cell_info["value"]).strip()

            for field, patterns in _ROSTER_COL_PATTERNS.items():
                for pattern in patterns:
                    if pattern.lower() == cell_str.lower() or pattern in cell_str:
                        if field not in detected:
                            detected[field] = cell_info["col"]  # 1-based
                        break

        # Need at least "name" column
        if "name" in detected:
            col_map = detected
            header_row_idx = row_idx
            break

    if col_map is None:
        # Fallback: standard layout B=No, C=class, D=name, ...
        warnings.append({
            "level": "warn", "child_name": None,
            "message": "名簿ヘッダー自動検出できず。標準レイアウト (B=No, D=氏名) を仮定します",
            "suggestion": None,
        })
        col_map = {
            "no": 2, "class_name": 3, "name": 4, "enrollment_type": 5,
            "child_order": 6, "birth_date": 7, "age_class": 8,
            "is_allergy": 9, "collection": 10,
        }
        header_row_idx = 4  # Assume row 5 is header (0-indexed)

    # ── Parse data rows ──
    empty_count = 0
    for row_idx in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[row_idx]

        # Build cell lookup by column
        cell_by_col = {}
        for cell_info in row:
            cell_by_col[cell_info["col"]] = cell_info["value"]

        def get_col(field):
            col = col_map.get(field)
            if col is None:
                return None
            return cell_by_col.get(col)

        # Read name
        name_val = get_col("name")
        if not name_val or not str(name_val).strip():
            empty_count += 1
            if empty_count >= 3:
                break  # Three consecutive empty = end of roster
            continue
        empty_count = 0

        name_str = str(name_val).strip()
        name_norm = normalize_name(name_str)

        # Parse each field
        roster_no = _safe_int(get_col("no"))
        class_name = str(get_col("class_name") or "").strip()

        enrollment_type = "月極"
        et_val = get_col("enrollment_type")
        if et_val:
            et_str = str(et_val).strip()
            if "一時" in et_str or "スポット" in et_str:
                enrollment_type = "一時"

        child_order = _safe_int(get_col("child_order")) or 1

        birth_date = _parse_date(get_col("birth_date"))

        age_class = None
        ac = get_col("age_class")
        if ac is not None:
            try:
                age_class = int(ac)
            except (ValueError, TypeError):
                m = re.match(r'(\d+)', str(ac))
                if m:
                    age_class = int(m.group(1))

        is_allergy = 0
        al_val = get_col("is_allergy")
        if al_val:
            al_str = str(al_val).strip()
            if al_str in ("〇", "○", "O", "o", "1", "TRUE", "true", "◯", "有", "あり"):
                is_allergy = 1

        collection_method = str(get_col("collection") or "口座振替").strip()

        # Get the Excel row number for this child
        excel_row = row[0]["row"] if row else row_idx + 1

        children.append({
            "roster_no": roster_no,
            "name": name_str,
            "name_norm": name_norm,
            "class_name": class_name,
            "enrollment_type": enrollment_type,
            "child_order": child_order,
            "birth_date": birth_date,
            "age_class": age_class,
            "is_allergy": is_allergy,
            "collection_method": collection_method,
            "row_index": excel_row,
        })

    if len(children) == 0:
        warnings.append({
            "level": "warn", "child_name": None,
            "message": "園児名簿からデータを抽出できませんでした",
            "suggestion": "名簿のレイアウトを確認してください",
        })

    return children, warnings


def _safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _parse_date(val) -> str | None:
    """Parse date to YYYY-MM-DD string"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()

    s = str(val).strip()
    if not s:
        return None

    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # Return as-is if can't parse
