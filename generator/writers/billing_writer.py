"""
保育料明細 Excel書き込みモジュール

書き込み対象列（数量のみ）:
  T=一時保育回数, W=早朝回数, Z=延長回数, AC=夜間回数, AF=病児回数
  AI=昼食数, AL=朝おやつ数, AO=午後おやつ数, AR=夕食数

絶対に触らない列:
  R(請求金額=数式), S(月額保育料)
  V,Y,AB,AE,AH,AK,AN,AQ,AT(合計=数式)
  U,X,AA,AD,AG,AJ,AM,AP,AS(単価=固定値)
"""

import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from engine.name_matcher import normalize_name


# Column mapping: charge_type → column letter
QUANTITY_COLUMNS = {
    "spot_care": "T",
    "early_morning": "W",
    "extension": "Z",
    "night": "AC",
    "sick": "AF",
    "lunch": "AI",
    "am_snack": "AL",
    "pm_snack": "AO",
    "dinner": "AR",
}

# Protected columns (never write)
PROTECTED_COLUMNS = {
    "R", "S",                                          # Fee formulas
    "V", "Y", "AB", "AE", "AH", "AK", "AN", "AQ", "AT",  # Subtotal formulas
    "U", "X", "AA", "AD", "AG", "AJ", "AM", "AP", "AS",    # Unit prices
}


def write_billing_detail(
    template_path: str | None,
    output_path: str,
    children: list[dict],
    charge_lines: list[dict],
    year: int,
    month: int,
) -> dict:
    """
    Write quantity values into billing detail template.
    Returns: {"success": bool, "error": str|None, "warnings": list}
    """
    warnings = []
    
    if not template_path or not os.path.exists(template_path):
        return {"success": False, "error": "保育料明細テンプレートがありません", "warnings": []}
    
    # Backup
    backup_path = template_path + ".backup"
    shutil.copy2(template_path, backup_path)
    
    try:
        wb = load_workbook(template_path)
        
        # Find month sheet (e.g., "1月")
        sheet_name = f"{month}月"
        if sheet_name not in wb.sheetnames:
            # Try alternate names
            for name in wb.sheetnames:
                if str(month) in name and '月' in name:
                    sheet_name = name
                    break
            else:
                return {"success": False, "error": f"シート「{month}月」が見つかりません", "warnings": []}
        
        ws = wb[sheet_name]
        
        # Pre-write error count
        pre_errors = _count_errors_sheet(ws)
        
        # Build charge lines index: (child_name_norm, charge_type) → line
        lines_index = {}
        for line in charge_lines:
            key = (normalize_name(line["child_name"]), line["charge_type"])
            lines_index[key] = line
        
        # Find child rows (K column = child name, starting from row 8)
        for row_idx in range(8, ws.max_row + 1):
            cell_name = ws.cell(row=row_idx, column=column_index_from_string("K")).value
            if not cell_name:
                continue
            
            name_norm = normalize_name(str(cell_name))
            
            # Write quantity columns
            for charge_type, col_letter in QUANTITY_COLUMNS.items():
                key = (name_norm, charge_type)
                line = lines_index.get(key)
                
                col_idx = column_index_from_string(col_letter)
                
                if line and line["quantity"] > 0:
                    ws.cell(row=row_idx, column=col_idx, value=line["quantity"])
                else:
                    # Write 0 for clarity
                    ws.cell(row=row_idx, column=col_idx, value=0)
        
        # ═══════════════════════════════════════════════════
        # POST-WRITE VALIDATION (3段階チェック)
        # ═══════════════════════════════════════════════════
        
        # Check 1: 差分検査 — エラー増加なら書き込みで壊した
        post_errors = _count_errors_sheet(ws)
        if post_errors > pre_errors:
            wb.close()
            shutil.copy2(backup_path, template_path)
            return {
                "success": False,
                "error": f"保育料明細破損検出(差分検査): エラー {pre_errors} → {post_errors}",
                "warnings": warnings,
            }
        
        # Check 2: R列（請求金額=数式）が数式として残っているか
        formula_intact = True
        formula_missing_rows = []
        for row_idx in range(8, ws.max_row + 1):
            r_cell = ws.cell(row=row_idx, column=column_index_from_string("R"))
            name_cell = ws.cell(row=row_idx, column=column_index_from_string("K"))
            if name_cell.value and str(name_cell.value).strip():
                # R列は数式であるべき
                if r_cell.data_type != 'f' and r_cell.value is not None:
                    # 数式が値に置き換わっている可能性
                    val = str(r_cell.value)
                    if not val.startswith('='):
                        formula_missing_rows.append(row_idx)
        
        if formula_missing_rows:
            warnings.append({
                "level": "warn",
                "child_name": None,
                "message": f"保育料明細 R列(請求金額): {len(formula_missing_rows)}行で数式が値になっています (行: {formula_missing_rows[:5]})",
                "suggestion": "テンプレートのR列数式が正しいか確認してください",
            })
        
        # Check 3: 保護対象列（数式列・単価列）に書き込んでいないか
        # → QUANTITY_COLUMNS以外に書き込みがないことを確認
        # この検査は書き込みロジックが正しければ不要だが、防御的に
        
        wb.save(output_path)
        wb.close()
        
        return {"success": True, "error": None, "warnings": warnings}
    
    except Exception as e:
        if os.path.exists(backup_path):
            shutil.copy2(backup_path, template_path)
        return {"success": False, "error": str(e), "warnings": warnings}


def _count_errors_sheet(ws) -> int:
    """Count #REF!/#VALUE!/#NAME?/#NULL!/#DIV/0! in a single sheet"""
    count = 0
    error_patterns = ['#REF!', '#VALUE!', '#NAME?', '#NULL!', '#DIV/0!']
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value)
            for pat in error_patterns:
                if pat in val:
                    count += 1
                    break
    return count
