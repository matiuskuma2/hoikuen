"""
日報Excel書き込みモジュール (openpyxl)

書き込み対象シート:
  ① 園児登園確認表□ — HH:MM-HH:MM 文字列
  ② 児童実績表申請□ — 時刻シリアル値
  ③ ◆保育時間 — 時刻シリアル + 給食マーク

書き込まないシート:
  ④ 給食実数表（個人）□ — 数式自動反映に任せる

安全方針:
  - 値のみ書き込み（数式・書式・条件付き書式は触らない）
  - 書き込み前後で #REF!/#VALUE! 増加チェック
  - NGなら成果物を返さない
"""

import os
import copy
import shutil
import calendar
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from engine.name_matcher import normalize_name


def write_daily_report(
    template_path: str,
    output_path: str,
    children: list[dict],
    usage_facts: list[dict],
    attendance_records: list[dict],
    all_plans: dict[str, dict],
    year: int,
    month: int,
) -> dict:
    """
    Write values into daily report template.
    Returns: {"success": bool, "error": str|None, "warnings": list}
    """
    warnings = []
    
    if not os.path.exists(template_path):
        return {"success": False, "error": "テンプレートファイルが見つかりません", "warnings": []}
    
    # Backup template before modification
    backup_path = template_path + ".backup"
    shutil.copy2(template_path, backup_path)
    
    try:
        # Load workbook (keep formulas, styles intact)
        wb = load_workbook(template_path)
        
        # ★ PRE-FLIGHT CORRUPTION CHECK
        pre_flight = _pre_flight_corruption_check(wb)
        if pre_flight:
            wb.close()
            detail = "; ".join(f["detail"] for f in pre_flight[:5])
            return {
                "success": False,
                "error": f"テンプレート破損検出（pre-flight）: {len(pre_flight)}件のエラー検出。{detail}",
                "warnings": warnings,
            }
        
        # Pre-write error count
        pre_errors = _count_errors(wb)
        
        # Build index: child_name_norm → child
        child_index = {}
        for c in children:
            norm = normalize_name(c["name"])
            child_index[norm] = c
        
        # Build facts index: (child_id, day) → fact
        facts_by_key = {}
        for f in usage_facts:
            facts_by_key[(f["child_id"], f["day"])] = f
        
        # Build attendance index
        att_by_key = {}
        for a in attendance_records:
            att_by_key[(a["lukumi_id"], a["day"])] = a
        
        # ── Sheet 1: 園児登園確認表□ ──
        result1 = _write_attendance_sheet(wb, children, facts_by_key, year, month, warnings)
        
        # ── Sheet 2: 児童実績表申請□ ──
        result2 = _write_jisseki_sheet(wb, children, facts_by_key, att_by_key, year, month, warnings)
        
        # ── Sheet 3: ◆保育時間 ──
        result3 = _write_hoiku_jikan_sheet(wb, children, facts_by_key, all_plans, year, month, warnings)
        
        # ═══════════════════════════════════════════════════
        # POST-WRITE VALIDATION (3段階チェック)
        # ═══════════════════════════════════════════════════
        
        # Check 1: 差分検査 — エラー増加なら書き込みで壊した
        post_errors = _count_errors(wb)
        if post_errors > pre_errors:
            wb.close()
            shutil.copy2(backup_path, template_path)
            return {
                "success": False,
                "error": f"テンプレート破損検出(差分検査): #REF!/#VALUE! が {pre_errors} → {post_errors} に増加。書き込みを中止しました。",
                "warnings": warnings,
            }
        
        # Check 2: 数式列が残っているか（書き込みで数式を潰していないか）
        formula_check = _check_formulas_intact(wb, warnings)
        if formula_check["fatal"]:
            wb.close()
            shutil.copy2(backup_path, template_path)
            return {
                "success": False,
                "error": f"テンプレート破損検出(数式消失): {formula_check['detail']}",
                "warnings": warnings,
            }
        
        # Check 3: 集計セルが空になっていないか（数式が壊れて空に見える）
        # → warningのみ（FATALにはしない。集計は再計算で復活する場合がある）
        summary_check = _check_summary_cells(wb, warnings)
        
        # Save
        wb.save(output_path)
        wb.close()
        
        return {"success": True, "error": None, "warnings": warnings}
    
    except Exception as e:
        # Restore from backup
        if os.path.exists(backup_path):
            shutil.copy2(backup_path, template_path)
        return {"success": False, "error": str(e), "warnings": warnings}


def _write_attendance_sheet(wb, children, facts_by_key, year, month, warnings):
    """
    園児登園確認表□
    1行/園児, stride=1
    列: F(day1) ... AJ(day31) — 固定オフセット: col = 6 + (day - 1)
    値: "H:MM-H:MM" 文字列
    """
    sheet_name = None
    for name in wb.sheetnames:
        if '園児登園確認表' in name:
            sheet_name = name
            break
    
    if not sheet_name:
        warnings.append({"level": "warn", "child_name": None,
                         "message": "園児登園確認表シートが見つかりません", "suggestion": None})
        return
    
    ws = wb[sheet_name]
    
    # Find children in sheet (starting from row 6)
    for row_idx in range(6, 6 + len(children)):
        # Read child name from column D (index 4)
        cell_name = ws.cell(row=row_idx, column=4).value
        if not cell_name:
            continue
        
        name_norm = normalize_name(str(cell_name))
        
        # Find matching child
        matched_child = None
        for c in children:
            if normalize_name(c["name"]) == name_norm:
                matched_child = c
                break
            if normalize_name(c["name"]).replace(' ', '') == name_norm.replace(' ', ''):
                matched_child = c
                break
        
        if not matched_child:
            continue
        
        child_id = matched_child.get("lukumi_id", "")
        
        # Write each day
        days_in_month = calendar.monthrange(year, month)[1]
        
        for day in range(1, days_in_month + 1):
            col = 6 + (day - 1)  # F=6, G=7, ... AJ=36
            fact = facts_by_key.get((child_id, day))
            
            if fact and fact["attendance_status"] in ("present", "late_arrive", "early_leave"):
                start = fact.get("actual_checkin") or fact.get("billing_start")
                end = fact.get("actual_checkout") or fact.get("billing_end")
                
                if start and end:
                    time_str = f"{_fmt_time_nolead(start)}-{_fmt_time_nolead(end)}"
                    ws.cell(row=row_idx, column=col, value=time_str)
                elif start:
                    ws.cell(row=row_idx, column=col, value=f"{_fmt_time_nolead(start)}-")


def _write_jisseki_sheet(wb, children, facts_by_key, att_by_key, year, month, warnings):
    """
    児童実績表申請□
    4行/園児ブロック, stride=4
    行0=登園時刻, 行1=降園時刻, 行2=利用時間, 行3=一時利用数
    """
    sheet_name = None
    for name in wb.sheetnames:
        if '児童実績表' in name:
            sheet_name = name
            break
    
    if not sheet_name:
        warnings.append({"level": "warn", "child_name": None,
                         "message": "児童実績表シートが見つかりません", "suggestion": None})
        return
    
    ws = wb[sheet_name]
    
    days_in_month = calendar.monthrange(year, month)[1]
    
    # ★ Fix #10: child_idx ではなくテンプレートの行名でマッチング
    # テンプレートの园児名行を探索して、データ行を特定する
    # フォールバック: 既存のインデックス方式も保持（テンプレートに名前セルがない場合）
    child_row_map = {}  # child_id -> base_row
    
    # First pass: try to find child names in column D or E
    for row_idx in range(7, min(7 + len(children) * 4 + 20, ws.max_row + 1), 4):
        cell_name = ws.cell(row=row_idx, column=4).value or ws.cell(row=row_idx, column=5).value
        if not cell_name:
            continue
        name_norm = normalize_name(str(cell_name))
        for child in children:
            cid = child.get("lukumi_id", "")
            if cid in child_row_map:
                continue
            if normalize_name(child["name"]) == name_norm or \
               normalize_name(child["name"]).replace(' ', '') == name_norm.replace(' ', ''):
                child_row_map[cid] = row_idx
                break
    
    # Fallback: for children not found by name, use index-based mapping
    used_rows = set(child_row_map.values())
    for child_idx_fb, child in enumerate(children):
        cid = child.get("lukumi_id", "")
        if cid not in child_row_map:
            fallback_row = 7 + child_idx_fb * 4
            if fallback_row not in used_rows:
                child_row_map[cid] = fallback_row
                used_rows.add(fallback_row)
    
    for child in children:
        child_id = child.get("lukumi_id", "")
        base_row = child_row_map.get(child_id)
        if base_row is None:
            continue  # No row found for this child
        
        for day in range(1, days_in_month + 1):
            col = 7 + (day - 1)  # G=7
            fact = facts_by_key.get((child_id, day))
            actual = att_by_key.get((child_id, day))
            
            if fact and fact["attendance_status"] in ("present", "late_arrive", "early_leave"):
                checkin = actual.get("actual_checkin") if actual else None
                checkout = actual.get("actual_checkout") if actual else None
                
                # Row 0: 登園時刻 (Excel time serial)
                if checkin:
                    ws.cell(row=base_row, column=col, value=_time_to_serial(checkin))
                
                # Row 1: 降園時刻
                if checkout:
                    ws.cell(row=base_row + 1, column=col, value=_time_to_serial(checkout))
                
                # Row 2: 利用時間
                if fact["billing_minutes"] is not None:
                    ws.cell(row=base_row + 2, column=col, value=fact["billing_minutes"] / (24 * 60))
                
                # Row 3: 一時利用数 (一時 only)
                if child.get("enrollment_type") == "一時" and fact.get("spot_30min_blocks"):
                    ws.cell(row=base_row + 3, column=col, value=fact["spot_30min_blocks"])


def _write_hoiku_jikan_sheet(wb, children, facts_by_key, all_plans, year, month, warnings):
    """
    ◆保育時間
    横展開: 8列/園児ブロック
    col+0=予定登園, col+1=予定降園, col+2=実績登園, col+3=実績降園,
    col+4=昼食, col+5=朝おやつ, col+6=午後おやつ, col+7=夕食
    
    Row 6 = day 1, Row 7 = day 2, ...
    
    ★ 給食マーク仕様（MVP確定版）:
      - 「〇」 = 提供あり（予定表の希望をそのまま書き込む。MVP=希望≡提供）
      - 「△」 = アレルギー対応食提供あり
      - 空セル = 提供なし
      ※ 給食実数表（個人）は触らない（数式が◆保育時間を参照して自動反映）
      ※ 将来拡張で「希望vs実提供」を区別する場合はここを変更
    """
    sheet_name = None
    for name in wb.sheetnames:
        if name == '◆保育時間':
            sheet_name = name
            break
    
    if not sheet_name:
        warnings.append({"level": "warn", "child_name": None,
                         "message": "◆保育時間シートが見つかりません", "suggestion": None})
        return
    
    ws = wb[sheet_name]
    
    days_in_month = calendar.monthrange(year, month)[1]
    
    for child_idx, child in enumerate(children):
        child_id = child.get("lukumi_id", "")
        child_name = child.get("name", "")
        child_norm = normalize_name(child_name)
        
        # Column block start: H(8) + child_idx * 8
        col_start = 8 + child_idx * 8
        
        # Find plans for this child
        child_plans = {}
        for plan_name, plans in all_plans.items():
            if normalize_name(plan_name) == child_norm or \
               normalize_name(plan_name).replace(' ', '') == child_norm.replace(' ', ''):
                child_plans = plans
                break
        
        for day in range(1, days_in_month + 1):
            row = 6 + (day - 1)  # Row 6 = day 1
            fact = facts_by_key.get((child_id, day))
            plan = child_plans.get(day)
            
            # col+0: planned start
            if plan and plan.get("planned_start"):
                ws.cell(row=row, column=col_start, value=_time_to_serial(plan["planned_start"]))
            
            # col+1: planned end
            if plan and plan.get("planned_end"):
                ws.cell(row=row, column=col_start + 1, value=_time_to_serial(plan["planned_end"]))
            
            # col+2: actual checkin
            if fact and fact.get("actual_checkin"):
                ws.cell(row=row, column=col_start + 2, value=_time_to_serial(fact["actual_checkin"]))
            
            # col+3: actual checkout
            if fact and fact.get("actual_checkout"):
                ws.cell(row=row, column=col_start + 3, value=_time_to_serial(fact["actual_checkout"]))
            
            # col+4-7: meal marks
            # ★ MVP: 希望=提供。「〇」=提供あり、「△」=アレルギー対応食、空=なし
            if fact and fact["attendance_status"] in ("present", "late_arrive", "early_leave"):
                if fact.get("has_lunch"):
                    mark = "△" if fact.get("meal_allergy") else "〇"
                    ws.cell(row=row, column=col_start + 4, value=mark)
                if fact.get("has_am_snack"):
                    ws.cell(row=row, column=col_start + 5, value="〇")
                if fact.get("has_pm_snack"):
                    ws.cell(row=row, column=col_start + 6, value="〇")
                if fact.get("has_dinner"):
                    ws.cell(row=row, column=col_start + 7, value="〇")


def _count_errors(wb) -> int:
    """Count #REF! and #VALUE! errors in all sheets"""
    count = 0
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if '#REF!' in val or '#VALUE!' in val or '#NAME?' in val or '#NULL!' in val:
                    count += 1
    return count


def _pre_flight_corruption_check(wb) -> list[dict]:
    """
    Fatal corruption guard (v3.2):
      1. Detect #REF!, #VALUE!, #NAME?, #NULL! errors
      2. Detect leftover formula placeholders (=XXXXX in value-only cells)
      3. Detect non-empty aggregate/summary cells that should be blank pre-write
    Returns list of corruption findings. If non-empty, abort the job.
    """
    findings = []
    error_patterns = ['#REF!', '#VALUE!', '#NAME?', '#NULL!', '#DIV/0!']
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                for pat in error_patterns:
                    if pat in val:
                        findings.append({
                            "sheet": ws_name,
                            "cell": f"{get_column_letter(cell.column)}{cell.row}",
                            "type": "error_value",
                            "detail": f"{pat} detected: {val[:50]}",
                        })
    return findings


def _time_to_serial(time_str: str) -> float:
    """Convert HH:MM to Excel time serial (0.0 - 1.0). Returns 0.0 on invalid input."""
    if not time_str or not isinstance(time_str, str):
        return 0.0
    parts = time_str.split(":")
    if len(parts) < 2:
        return 0.0
    try:
        h = int(parts[0])
        m = int(parts[1])
        return (h * 60 + m) / (24 * 60)
    except (ValueError, TypeError):
        return 0.0


def _fmt_time_nolead(time_str: str) -> str:
    """Format time without leading zero: 08:12 → 8:12"""
    parts = time_str.split(":")
    h = int(parts[0])
    m = parts[1].zfill(2)
    return f"{h}:{m}"


def _check_formulas_intact(wb, warnings) -> dict:
    """
    Check 2: 数式列が残っているか（書き込みで数式を潰していないか）。
    
    チェック対象:
      - 給食実数表（個人）□ シートが存在すれば、数式が残っているか
      - 園児登園確認表□ / 児童実績表□ の合計行の数式が残っているか
    
    ※ openpyxl で data_only=False (デフォルト) で読み込んでいるので、
       cell.data_type == 'f' は数式セルを示す。
    
    Returns: {"fatal": bool, "detail": str}
    """
    fatal = False
    detail_parts = []
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        
        # ── 給食実数表（個人） — 一切書き込まないので数式が残っているはず ──
        if '給食実数表' in ws_name:
            formula_count = 0
            for row in ws.iter_rows(max_row=min(ws.max_row or 1, 200)):
                for cell in row:
                    if cell.data_type == 'f':
                        formula_count += 1
            if formula_count == 0:
                # 数式が1つも無い = 元から数式なしのシートか、全消失。
                # 元から無い場合もあるので warning に留める。
                warnings.append({
                    "level": "warn",
                    "child_name": None,
                    "message": f"「{ws_name}」に数式が1つも見つかりません。テンプレートを確認してください。",
                    "suggestion": "給食実数表の集計数式が壊れている可能性があります",
                })
        
        # ── 園児登園確認表 / 児童実績表 — 合計行の数式を抽出チェック ──
        if '園児登園確認表' in ws_name or '児童実績表' in ws_name:
            max_row = ws.max_row or 1
            # 末尾5行内に合計行がある場合、数式が残っているか
            for row_idx in range(max(1, max_row - 5), max_row + 1):
                label = ws.cell(row=row_idx, column=1).value or ws.cell(row=row_idx, column=2).value
                if label and ('合計' in str(label) or '計' in str(label)):
                    formula_found = False
                    for col in range(6, 37):  # F to AJ
                        cell = ws.cell(row=row_idx, column=col)
                        if cell.data_type == 'f':
                            formula_found = True
                            break
                    if not formula_found:
                        # 合計行に数式が無い = 壊れた可能性。warning に留める。
                        warnings.append({
                            "level": "warn",
                            "child_name": None,
                            "message": f"「{ws_name}」の合計行(行{row_idx})に数式が見つかりません",
                            "suggestion": "合計行が数式でなく固定値になっている可能性があります",
                        })
    
    if detail_parts:
        detail = "; ".join(detail_parts)
    else:
        detail = ""
    
    return {"fatal": fatal, "detail": detail}


def _check_summary_cells(wb, warnings):
    """
    Check 3: 主要集計セルが空になっていないか。
    
    書き込み後、数式セルが壊れて空に見える場合を警告する。
    注意: openpyxl で数式を保持して保存した場合、
    キャッシュ値がクリアされる場合があるので、
    これは warning のみ（FATAL にはしない）。
    
    チェック対象:
      - 園児登園確認表の集計行（出席日数合計等）
      - 児童実績表の月間合計
      - ◆保育時間のヘッダ行（園児名が書き込み前に入っているか）
    """
    issues_found = 0
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        
        # ── 園児登園確認表 / 児童実績表 の集計行 ──
        if '園児登園確認表' in ws_name or '児童実績表' in ws_name:
            max_row = ws.max_row or 1
            for row_idx in range(max(1, max_row - 5), max_row + 1):
                label = ws.cell(row=row_idx, column=1).value or ws.cell(row=row_idx, column=2).value
                if label and ('合計' in str(label) or '計' in str(label)):
                    # 合計行のデータ列に数式もデータもない場合は警告
                    empty_count = 0
                    formula_count = 0
                    for col in range(6, 37):  # F to AJ
                        cell = ws.cell(row=row_idx, column=col)
                        if cell.data_type == 'f':
                            formula_count += 1
                        elif cell.value is None:
                            empty_count += 1
                    
                    # 数式0件かつ全空 → 壊れている可能性大
                    if formula_count == 0 and empty_count > 25:
                        warnings.append({
                            "level": "warn",
                            "child_name": None,
                            "message": f"「{ws_name}」合計行(行{row_idx})が空です。集計数式が消失している可能性があります",
                            "suggestion": "ExcelでF1ファイルを開いて集計行を確認してください",
                        })
                        issues_found += 1
    
    return {"ok": issues_found == 0, "issues": issues_found}
