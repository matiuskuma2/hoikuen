"""
Create test data files for end-to-end testing of the Ayukko pipeline.
Generates:
  1. Lukumi attendance Excel (mock 66 children x 28 days)
  2. Schedule plan Excel files (3 children)
  3. Daily report template (with sheets: 園児登園確認表, 児童実績表, ◆保育時間, ◆園児名簿)
  4. Billing detail template
"""

import os
import random
from datetime import datetime, date, time
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
YEAR = 2026
MONTH = 2

# ── Mock children data ──
CHILDREN = [
    {"id": "1001", "surname": "田中", "firstname": "太郎", "kana_sei": "タナカ", "kana_mei": "タロウ", "birth": "2023-04-15", "age": 2, "class": "ぞう組"},
    {"id": "1002", "surname": "佐藤", "firstname": "花子", "kana_sei": "サトウ", "kana_mei": "ハナコ", "birth": "2022-08-20", "age": 3, "class": "きりん組"},
    {"id": "1003", "surname": "鈴木", "firstname": "一郎", "kana_sei": "スズキ", "kana_mei": "イチロウ", "birth": "2021-12-01", "age": 4, "class": "らいおん組"},
    {"id": "1004", "surname": "高橋", "firstname": "美咲", "kana_sei": "タカハシ", "kana_mei": "ミサキ", "birth": "2024-01-10", "age": 1, "class": "うさぎ組"},
    {"id": "1005", "surname": "Mondal", "firstname": "Aum", "kana_sei": "モンダル", "kana_mei": "アウム", "birth": "2025-02-06", "age": 0, "class": "ひよこ組"},
]


def create_lukumi_excel():
    """Create mock Lukumi attendance data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "登降園データ"

    # Header row
    headers = ["クラス名", "園児姓", "園児名", "日付", "登園日時", "降園日時",
               "メモ", "園児ID", "姓よみ", "名よみ", "生年月日", "クラス年齢"]
    ws.append(headers)

    import calendar
    days_in_month = calendar.monthrange(YEAR, MONTH)[1]

    for child in CHILDREN:
        for day in range(1, days_in_month + 1):
            dt = date(YEAR, MONTH, day)
            # Skip weekends roughly
            if dt.weekday() >= 5:
                continue

            # Random attendance
            if random.random() < 0.1:
                continue  # 10% absent

            checkin_h = random.choice([7, 8, 8, 8, 9, 9])
            checkin_m = random.randint(0, 59)
            checkout_h = random.choice([15, 16, 17, 17, 18, 18])
            checkout_m = random.randint(0, 59)

            checkin_dt = datetime(YEAR, MONTH, day, checkin_h, checkin_m, 0)
            checkout_dt = datetime(YEAR, MONTH, day, checkout_h, checkout_m, 0)

            row = [
                child["class"],
                child["surname"],
                child["firstname"],
                dt,
                checkin_dt,
                checkout_dt,
                "",
                child["id"],
                child["kana_sei"],
                child["kana_mei"],
                child["birth"],
                child["age"],
            ]
            ws.append(row)

    path = os.path.join(OUTPUT_DIR, "lukumi_test.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def create_schedule_excel(child):
    """Create a single child's schedule plan."""
    wb = Workbook()
    ws = wb.active
    ws.title = "原本"

    # B6 = child name
    ws.cell(row=6, column=2, value=f"{child['surname']} {child['firstname']}")

    # J1 = year, M1 = month
    ws.cell(row=1, column=10, value=YEAR)
    ws.cell(row=1, column=13, value=MONTH)

    import calendar
    days_in_month = calendar.monthrange(YEAR, MONTH)[1]

    # Left half: days 1-15
    for i in range(min(15, days_in_month)):
        day = i + 1
        row = 12 + i
        dt = date(YEAR, MONTH, day)

        ws.cell(row=row, column=2, value=day)  # B=date

        if dt.weekday() < 5:  # weekdays only
            start_h = random.choice([8, 9])
            start_m = random.choice([0, 30])
            end_h = random.choice([15, 16, 17])
            end_m = random.choice([0, 30])

            ws.cell(row=row, column=4, value=time(start_h, start_m))  # D=start
            ws.cell(row=row, column=7, value=time(end_h, end_m))     # G=end
            ws.cell(row=row, column=10, value="〇")                  # J=lunch
            ws.cell(row=row, column=11, value="〇")                  # K=snack
            if end_h >= 17:
                ws.cell(row=row, column=12, value="〇")              # L=dinner

    # Right half: days 16-31
    for i in range(days_in_month - 15):
        day = 16 + i
        row = 12 + i
        dt = date(YEAR, MONTH, day)

        ws.cell(row=row, column=13, value=day)  # M=date

        if dt.weekday() < 5:
            start_h = random.choice([8, 9])
            start_m = random.choice([0, 30])
            end_h = random.choice([15, 16, 17])
            end_m = random.choice([0, 30])

            ws.cell(row=row, column=15, value=time(start_h, start_m))  # O=start
            ws.cell(row=row, column=18, value=time(end_h, end_m))     # R=end
            ws.cell(row=row, column=21, value="〇")                  # U=lunch
            ws.cell(row=row, column=22, value="〇")                  # V=snack

    name = f"{child['surname']}_{child['firstname']}"
    path = os.path.join(OUTPUT_DIR, f"schedule_{name}.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def create_daily_report_template():
    """Create a daily report template with required sheets."""
    wb = Workbook()

    # Sheet 1: 園児登園確認表
    ws1 = wb.active
    ws1.title = "園児登園確認表□"
    ws1.cell(row=1, column=1, value="園児登園確認表")
    ws1.cell(row=3, column=1, value=f"{YEAR}年{MONTH}月")

    # Header row 5: day columns
    ws1.cell(row=5, column=4, value="氏名")
    import calendar
    days_in_month = calendar.monthrange(YEAR, MONTH)[1]
    for d in range(1, days_in_month + 1):
        ws1.cell(row=5, column=6 + (d - 1), value=d)

    # Child rows starting at row 6
    for i, child in enumerate(CHILDREN):
        ws1.cell(row=6 + i, column=4, value=f"{child['surname']} {child['firstname']}")

    # Sheet 2: 児童実績表申請
    ws2 = wb.create_sheet("児童実績表申請□")
    ws2.cell(row=1, column=1, value="児童実績表")
    for i, child in enumerate(CHILDREN):
        base_row = 7 + i * 4
        ws2.cell(row=base_row, column=1, value=f"{child['surname']} {child['firstname']}")
        ws2.cell(row=base_row, column=2, value="登園")
        ws2.cell(row=base_row + 1, column=2, value="降園")
        ws2.cell(row=base_row + 2, column=2, value="時間")
        ws2.cell(row=base_row + 3, column=2, value="一時")

    # Sheet 3: ◆保育時間
    ws3 = wb.create_sheet("◆保育時間")
    ws3.cell(row=1, column=1, value="保育時間一覧")
    for i, child in enumerate(CHILDREN):
        col_start = 8 + i * 8
        ws3.cell(row=4, column=col_start, value=f"{child['surname']} {child['firstname']}")
        ws3.cell(row=5, column=col_start, value="予定登")
        ws3.cell(row=5, column=col_start + 1, value="予定降")
        ws3.cell(row=5, column=col_start + 2, value="実績登")
        ws3.cell(row=5, column=col_start + 3, value="実績降")
        ws3.cell(row=5, column=col_start + 4, value="昼食")
        ws3.cell(row=5, column=col_start + 5, value="朝お")
        ws3.cell(row=5, column=col_start + 6, value="午お")
        ws3.cell(row=5, column=col_start + 7, value="夕食")

    # Sheet 4: ◆園児名簿
    ws4 = wb.create_sheet("◆園児名簿")
    ws4.cell(row=5, column=2, value="No")
    ws4.cell(row=5, column=3, value="クラス名")
    ws4.cell(row=5, column=4, value="氏名")
    ws4.cell(row=5, column=5, value="利用区分")
    ws4.cell(row=5, column=6, value="第何子")
    ws4.cell(row=5, column=7, value="生年月日")
    ws4.cell(row=5, column=8, value="歳児")
    ws4.cell(row=5, column=9, value="アレルギー")
    ws4.cell(row=5, column=10, value="徴収方法")

    for i, child in enumerate(CHILDREN):
        row = 6 + i
        ws4.cell(row=row, column=2, value=i + 1)
        ws4.cell(row=row, column=3, value=child["class"])
        ws4.cell(row=row, column=4, value=f"{child['surname']} {child['firstname']}")
        ws4.cell(row=row, column=5, value="月極")
        ws4.cell(row=row, column=6, value=1)
        ws4.cell(row=row, column=7, value=child["birth"])
        ws4.cell(row=row, column=8, value=child["age"])
        ws4.cell(row=row, column=9, value="")
        ws4.cell(row=row, column=10, value="口座振替")

    path = os.path.join(OUTPUT_DIR, "daily_report_template.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def create_billing_template():
    """Create a billing detail template."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{MONTH}月"

    # Header area
    ws.cell(row=1, column=1, value=f"保育料明細書 {YEAR}年{MONTH}月")

    # Column headers at row 7
    from openpyxl.utils import column_index_from_string
    headers = {
        "K": "園児名",
        "R": "請求金額",
        "S": "月額保育料",
        "T": "一時保育",
        "U": "単価",
        "V": "小計",
        "W": "早朝",
        "X": "単価",
        "Y": "小計",
        "Z": "延長",
        "AA": "単価",
        "AB": "小計",
        "AC": "夜間",
        "AD": "単価",
        "AE": "小計",
        "AF": "病児",
        "AG": "単価",
        "AH": "小計",
        "AI": "昼食",
        "AJ": "単価",
        "AK": "小計",
        "AL": "朝おやつ",
        "AM": "単価",
        "AN": "小計",
        "AO": "午後おやつ",
        "AP": "単価",
        "AQ": "小計",
        "AR": "夕食",
        "AS": "単価",
        "AT": "小計",
    }

    for col_letter, header_text in headers.items():
        col_idx = column_index_from_string(col_letter)
        ws.cell(row=7, column=col_idx, value=header_text)

    # Child rows starting at row 8
    for i, child in enumerate(CHILDREN):
        row = 8 + i
        ws.cell(row=row, column=column_index_from_string("K"),
                value=f"{child['surname']} {child['firstname']}")
        # Formula placeholders for R column
        ws.cell(row=row, column=column_index_from_string("R"),
                value=f"=S{row}+V{row}+Y{row}+AB{row}+AE{row}+AH{row}+AK{row}+AN{row}+AQ{row}+AT{row}")

    path = os.path.join(OUTPUT_DIR, "billing_template.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


if __name__ == "__main__":
    random.seed(42)  # Reproducible
    print(f"=== Creating test data for {YEAR}/{MONTH:02d} ===")
    create_lukumi_excel()
    for child in CHILDREN[:3]:  # Create schedules for first 3 children
        create_schedule_excel(child)
    create_daily_report_template()
    create_billing_template()
    print("=== Done ===")
